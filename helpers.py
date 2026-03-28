import csv
import io
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from constants import (
    BADGE_LOGO_FILE,
    HEADER_LOGO_FILE,
    WARNING_DAYS,
)


def format_date(value):
    if not value:
        return "N/A"
    return value.strftime("%Y-%m-%d")


def format_number(value, decimals=0):
    if value is None or value == "":
        return "N/A"

    try:
        value = float(value)
    except (TypeError, ValueError):
        return str(value)

    if decimals == 0:
        if value.is_integer():
            return str(int(value))
        return f"{value:.0f}"

    return f"{value:.{decimals}f}"


def format_hours(hours):
    if hours is None:
        return "N/A"

    try:
        hours = float(hours)
    except (TypeError, ValueError):
        return "N/A"

    if hours < 0:
        return "0h"

    return f"{int(hours)}h"


def format_days(days):
    if days is None:
        return "N/A"

    try:
        days = float(days)
    except (TypeError, ValueError):
        return "N/A"

    if days < 0:
        return "0 days"

    return f"{int(days)} days"


def component_status(component):
    remaining_hours = component.get("remaining_hours")

    if remaining_hours is None:
        return "Warning"

    try:
        remaining_hours = float(remaining_hours)
    except (TypeError, ValueError):
        return "Warning"

    if remaining_hours < 0:
        return "Out of service"

    if remaining_hours <= WARNING_DAYS * 24:
        return "Warning"

    return "Normal"


def system_status(system):
    remarks = str(system.get("remarks", "") or "").strip().lower()
    if "not-in-service" in remarks or "out of service" in remarks:
        return "Out of service"

    components = system.get("components", [])
    if components:
        statuses = [component_status(component) for component in components]

        if "Out of service" in statuses:
            return "Out of service"
        if "Warning" in statuses:
            return "Warning"
        return "Normal"

    remaining_hours = system.get("remaining_hours")
    remaining_days = system.get("remaining_days")

    if remaining_hours is None or remaining_days is None:
        return "Warning"

    try:
        remaining_hours = float(remaining_hours)
        remaining_days = float(remaining_days)
    except (TypeError, ValueError):
        return "Warning"

    if remaining_hours < 0:
        return "Out of service"

    if remaining_days <= WARNING_DAYS:
        return "Warning"

    return "Normal"


def overdue_badge(_remaining_hours):
    return ""


def status_badge(status):
    normalized = str(status).strip().lower()

    if normalized == "normal":
        return '<span class="badge badge-ok">Normal</span>'
    if normalized == "warning":
        return '<span class="badge badge-warning">Warning</span>'
    if normalized == "out of service":
        return '<span class="badge badge-oos">Out of service</span>'

    return f'<span class="badge badge-neutral">{status}</span>'


def room_colors(status):
    base_background = "linear-gradient(180deg, #ffffff 0%, #f8fafc 100%)"
    neutral_border = "#d8e0ea"

    if status == "Normal":
        return {
            "background": base_background,
            "border": neutral_border,
            "title": "#0f172a",
            "badge": '<span class="badge badge-ok">Normal</span>',
            "accent": "#16a34a",
        }

    if status == "Warning":
        return {
            "background": base_background,
            "border": neutral_border,
            "title": "#0f172a",
            "badge": '<span class="badge badge-warning">Warning</span>',
            "accent": "#d97706",
        }

    return {
        "background": base_background,
        "border": neutral_border,
        "title": "#0f172a",
        "badge": '<span class="badge badge-oos">Out of service</span>',
        "accent": "#dc2626",
    }


def build_remarks_text(system):
    remark = str(system.get("remarks", "") or "").strip()

    if remark.lower() in {"n/a", "na", "-", "?"}:
        return ""

    return remark


def matches_search(system, search_text):
    if not search_text:
        return True

    query = search_text.lower().strip()

    searchable = " ".join(
        [
            str(system.get("building", "")),
            str(system.get("location", "")),
            str(system.get("model", "")),
            build_remarks_text(system),
            str(system.get("power_rating_label", "")),
        ]
    ).lower()

    return query in searchable


def system_sort_key(system, mode):
    status = system_status(system)
    status_priority = {
        "Out of service": 0,
        "Warning": 1,
        "Normal": 2,
    }

    building = str(system.get("building", "") or "")
    location = str(system.get("location", "") or "")
    remaining_days = system.get("remaining_days")

    try:
        remaining_days_sort = float(remaining_days) if remaining_days is not None else 999999
    except (TypeError, ValueError):
        remaining_days_sort = 999999

    start_date = system.get("start_date")

    if mode == "Urgency":
        return (status_priority.get(status, 9), remaining_days_sort, building.lower(), location.lower())

    if mode == "Remaining days":
        return (remaining_days_sort, status_priority.get(status, 9), building.lower(), location.lower())

    if mode == "Start date":
        return (start_date is None, start_date or 0, building.lower(), location.lower())

    if mode == "Building":
        return (building.lower(), location.lower())

    if mode == "Location":
        return (location.lower(), building.lower())

    return (status_priority.get(status, 9), remaining_days_sort, building.lower(), location.lower())


def kpi_summary(systems):
    statuses = [system_status(system) for system in systems]

    return {
        "total_systems": len(systems),
        "normal_systems": sum(1 for status in statuses if status == "Normal"),
        "warning_systems": sum(1 for status in statuses if status == "Warning"),
        "out_of_service_systems": sum(1 for status in statuses if status == "Out of service"),
        "total_lamps": sum(int(float(system.get("number_of_lamps") or 0)) for system in systems),
    }


def export_csv_bytes(systems):
    output = io.StringIO()

    fieldnames = [
        "system_id",
        "building",
        "location",
        "status",
        "model",
        "actual_flow_rate",
        "current_dosing",
        "upgrade_dosing",
        "connection_size_mm",
        "number_of_lamps",
        "design_flow_rate",
        "power_rating_label",
        "start_date",
        "end_date",
        "remaining_days",
        "remaining_hours",
        "remarks",
    ]

    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    for system in systems:
        writer.writerow(
            {
                "system_id": system.get("system_id"),
                "building": system.get("building"),
                "location": system.get("location"),
                "status": system_status(system),
                "model": system.get("model"),
                "actual_flow_rate": system.get("actual_flow_rate"),
                "current_dosing": system.get("current_dosing"),
                "upgrade_dosing": system.get("upgrade_dosing"),
                "connection_size_mm": system.get("connection_size_mm"),
                "number_of_lamps": system.get("number_of_lamps"),
                "design_flow_rate": system.get("design_flow_rate"),
                "power_rating_label": system.get("power_rating_label"),
                "start_date": format_date(system.get("start_date")),
                "end_date": format_date(system.get("end_date")),
                "remaining_days": system.get("remaining_days"),
                "remaining_hours": system.get("remaining_hours"),
                "remarks": build_remarks_text(system),
            }
        )

    return output.getvalue().encode("utf-8")


def export_excel_bytes(data, systems):
    meta = data.get("meta", {})

    wb = Workbook()
    ws = wb.active
    ws.title = meta.get("sheet_name") or "UV LIST Rev. 01"

    if BADGE_LOGO_FILE.exists():
        img = XLImage(str(BADGE_LOGO_FILE))
        img.width = 96
        img.height = 84
        ws.add_image(img, "A1")

    if HEADER_LOGO_FILE.exists():
        img = XLImage(str(HEADER_LOGO_FILE))
        img.width = 442
        img.height = 108
        ws.add_image(img, "T1")

    ws["I1"] = meta.get("organization") or "AQUATIC RESEARCH CENTER"
    ws["I2"] = meta.get("department") or "MAINTENANCE DEPARTMENT"
    ws["I3"] = meta.get("title") or "UV STERILIZER SYSTEM MONITORING"
    ws["Q4"] = "MAX DAYS"
    ws["R4"] = meta.get("max_days") or 625
    ws["S4"] = "MAX HOURS"
    ws["T4"] = meta.get("max_hours") or 15000

    headers = [
        "No.",
        "BUILDING",
        "LOCATION",
        "MAKE",
        "MODEL",
        "Actual Flow Rate (m3/hr)",
        "CURRENT DOSING (mJ/cm2)",
        "UPGRADE DOSING (mJ/cm2)",
        "Connection Size (mm)",
        "NUMBER OF LAMPS",
        "DESIGN FLOW RATE M3/H",
        "LAMP REPLACEMENT",
        "BALLAST REPLACEMENT",
        "ELECTRICAL POWER KW/LAMP",
        "",
        "START DATE",
        "END DATE",
        "REMAINING DAYS",
        "REMAINING HOURS",
        "REMAINING DAYS\n(BAR GRAPH)",
        "REMARKS",
        "",
    ]

    for col_index, header in enumerate(headers, start=1):
        ws.cell(5, col_index).value = header

    thin_gray = Side(style="thin", color="D9E2EC")
    header_fill = PatternFill("solid", fgColor="0E4F2F")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    title_font = Font(bold=True, size=16, color="0F172A")
    subtitle_font = Font(bold=True, size=13, color="0F172A")
    section_font = Font(bold=True, size=11, color="0F172A")

    ws["I1"].font = subtitle_font
    ws["I2"].font = subtitle_font
    ws["I3"].font = title_font
    ws["Q4"].font = section_font
    ws["S4"].font = section_font

    for col_index in range(1, 23):
        cell = ws.cell(5, col_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin_gray, bottom=thin_gray)

    row_index = 6
    for item_index, system in enumerate(systems, start=1):
        remarks = build_remarks_text(system)

        row_values = [
            item_index,
            system.get("building"),
            system.get("location"),
            "",
            system.get("model"),
            system.get("actual_flow_rate"),
            system.get("current_dosing"),
            system.get("upgrade_dosing"),
            system.get("connection_size_mm"),
            int(float(system.get("number_of_lamps") or 0)),
            system.get("design_flow_rate"),
            system.get("lamp_replacement"),
            system.get("ballast_replacement"),
            system.get("power_rating_label"),
            "",
            system.get("start_date"),
            system.get("end_date"),
            system.get("remaining_days"),
            system.get("remaining_hours"),
            max(float(system.get("remaining_days") or 0), 0),
            remarks,
            "",
        ]

        for col_index, value in enumerate(row_values, start=1):
            cell = ws.cell(row_index, col_index)
            cell.value = value
            cell.fill = white_fill
            cell.border = Border(bottom=thin_gray)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.cell(row_index, 16).number_format = "yyyy-mm-dd"
        ws.cell(row_index, 17).number_format = "yyyy-mm-dd"
        row_index += 1

    ws.conditional_formatting.add(
        f"T6:T{max(row_index - 1, 6)}",
        DataBarRule(
            start_type="num",
            start_value=0,
            end_type="max",
            color="5B9BD5",
            showValue=False,
        ),
    )

    widths = {
        "A": 7,
        "B": 15,
        "C": 28,
        "D": 12,
        "E": 18,
        "F": 18,
        "G": 20,
        "H": 20,
        "I": 18,
        "J": 16,
        "K": 20,
        "L": 24,
        "M": 24,
        "N": 18,
        "O": 4,
        "P": 14,
        "Q": 14,
        "R": 15,
        "S": 16,
        "T": 18,
        "U": 40,
        "V": 4,
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[5].height = 34

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()


def find_system_by_id(systems, system_id):
    for system in systems:
        if str(system.get("system_id")) == str(system_id):
            return system
    return None


def find_system_index_by_id(systems, system_id):
    for index, system in enumerate(systems):
        if str(system.get("system_id")) == str(system_id):
            return index
    return None